from openpyxl import load_workbook
from django.contrib.auth.hashers import make_password
from datetime import datetime
from .models import Student


def import_students(file):

    workbook = load_workbook(file, read_only=True)
    sheet = workbook.active

    count = 0

    for row in sheet.iter_rows(min_row=2, values_only=True):

        if not row[0]:
            continue

        roll = str(row[0]).strip()
        name = str(row[1]).strip()

        password_value = row[2]

        # If the password is a date, convert it to DD-MM-YYYY
        if isinstance(password_value, datetime):
            password = password_value.strftime("%d-%m-%Y")

        # If it's a number (phone number or numeric password)
        elif isinstance(password_value, (int, float)):
            password = str(int(password_value))

        # Otherwise treat it as text
        else:
            password = str(password_value).strip()

        Student.objects.update_or_create(
            roll_number=roll,
            defaults={
                "name": name,
                "password": make_password(password),
            },
        )

        count += 1

    return count